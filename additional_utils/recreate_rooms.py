import os
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from modules.consts.common import OUTPUT_DIR_PATH
from modules.reader import read_all_info
from modules.consts.common import WORKBOOK_HEADER_PRIVATE
from modules.consts.common import WORKBOOK_HEADER_PUBLIC
from modules.time_manager import converter_date
from modules.time_manager import converter_time
from modules.consts.common import INPUT_DIR_PATH
from modules.consts.common import HEADERS, ROOMS
import requests
from modules.writer import formating_data


def read():
    all_data = []
    id_to_books = {}
    files = os.listdir(INPUT_DIR_PATH)
    exel_files = filter(lambda x: x.endswith(".xlsx"), files)
    for file in exel_files:
        data_from_book = no_full_read_exel_file(file)
        all_data = [*all_data, *data_from_book]
    return all_data


def no_full_read_exel_file(file):
    book_data = []
    row = 2
    path = os.path.join(INPUT_DIR_PATH, file)
    while True:
        wb = load_workbook(path, data_only=True)
        sheet = wb["Sheet"]
        event_id = sheet.cell(row=row, column=12).value
        book_data.append(event_id)
        row += 1
        if sheet.cell(row=row, column=1).value is None:
            break
    return book_data


def get_another_v(event_id, user_id):
    url = f"https://userapi.webinar.ru/v3/organization/events/{event_id}/move"
    body = {"userId": user_id}
    requests.put(url, data=body, headers=HEADERS)
    print("ok")


book_data = read()

OUTPUT_FILE_PATH = os.path.join(OUTPUT_DIR_PATH, "svodniy_table.xlsx")
if "svodniy_table.xlsx" not in os.listdir(OUTPUT_DIR_PATH):
    work_book = Workbook()
    work_book.save(filename=str(OUTPUT_FILE_PATH))

work_book = load_workbook(OUTPUT_FILE_PATH)
sheet_names = work_book.sheetnames
sheet = work_book[sheet_names[0]]
id = 0

for row, event_id in enumerate(book_data):
    sheet.cell(row=int(row) + 2, column=int(10)).value = "w{}".format(ROOMS[id][1])
    # print(f'https://userapi.webinar.ru/v3/organization/events/{event_id}/move')
    # get_another_v(event_id, ROOMS[id][0])
    id += 1
    if len(ROOMS) == id:
        id = 0

work_book.save(filename=str(OUTPUT_FILE_PATH))


def get_another_v(event_id, user_id):
    url = f"https://userapi.webinar.ru/v3/organization/events/{event_id}/move"
    body = {userId: user_id}
    responce = requests.put(url, data=body, headers=HEADERS).json()
    print(responce)
