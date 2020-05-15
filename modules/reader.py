from modules.time_manager import converter_date
from modules.time_manager import converter_time
from openpyxl import load_workbook
from .consts.common import INPUT_DIR_PATH
import os


def read_row_from_exel(sheet, row):
    start_t = str(sheet.cell(row=row, column=7).value)
    CELLS = {
        "date": converter_date(sheet.cell(row=row, column=1).value),
        "id": str(sheet.cell(row=row, column=2).value),
        "name": sheet.cell(row=row, column=3).value,
        "email": sheet.cell(row=row, column=4).value,
        "phone_number": sheet.cell(row=row, column=5).value,
        "subject": sheet.cell(row=row, column=6).value,
        "start_t": converter_time(time=start_t),
        "end_time": str(sheet.cell(row=row, column=8).value),
        "group": str(sheet.cell(row=row, column=11).value),
    }
    output_data = (CELLS, start_t)
    return output_data


def read_email_from_parametrs(wb):
    try:
        return str(wb["Параметры"].cell(row=6, column=2).value)
    except KeyError:
        return 'zhuplev@gmail.com'


def full_read_exel_file(file, id_to_books):
    book_data = []
    row = 2
    path = os.path.join(INPUT_DIR_PATH, file)
    while True:
        wb = load_workbook(path, data_only=True)
        sheet = wb["Вебинары"]
        data_from_row = read_row_from_exel(sheet, row)
        book_data.append(data_from_row)
        id_to_books[data_from_row[0]["id"]] = file
        row += 1
        if sheet.cell(row=row, column=1).value is None:
            break
    email = read_email_from_parametrs(wb=wb)
    return book_data, id_to_books, email


def read_all_info():
    all_data = []
    id_to_books = {}
    file_to_email = {}
    files = [prepair_file_name(file) for file in os.listdir(INPUT_DIR_PATH)]
    exel_files = filter(lambda x: x.endswith(".xlsx"), files)
    for file in exel_files:
        data_from_book, id_to_books, email = full_read_exel_file(file, id_to_books)
        all_data = [*all_data, *data_from_book]
        file_to_email[file] = email
    return all_data, id_to_books, file_to_email


def prepair_file_name(file_name):
    name = ".".join(
        [".".join(file_name.split(".")[:-1]), file_name.split(".")[-1].lower()]
    )
    return name
