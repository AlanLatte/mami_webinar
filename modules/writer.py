import os
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from .consts.common import OUTPUT_DIR_PATH

from modules.consts.common import WORKBOOK_HEADER_PRIVATE
from modules.consts.common import WORKBOOK_HEADER_PUBLIC


def write_data(sheet, wb, event_session_id, link, room, row, path):
    sheet.cell(row=row, column=9).value = link
    sheet.cell(row=row, column=10).value = room
    sheet.cell(row=row, column=15).value = event_session_id
    wb.save(path)


def create_workbook(data: list, name: str, params: dict):
    if params["type"] == "private":
        HEADER = WORKBOOK_HEADER_PRIVATE

    elif params["type"] == "public":
        HEADER = WORKBOOK_HEADER_PUBLIC
    else:
        print("create_workbook need some params!")
        sys.exit()

    OUTPUT_FILE_PATH = os.path.join(OUTPUT_DIR_PATH, name)

    data = formating_data(data, params)

    if name not in os.listdir(OUTPUT_DIR_PATH):
        work_book = Workbook()
        work_book.save(filename=str(OUTPUT_FILE_PATH))

    work_book = load_workbook(OUTPUT_FILE_PATH)
    sheet_names = work_book.sheetnames
    if sheet_names.__len__() is 1:
        sheet = work_book[sheet_names[0]]
        for column, title_object in enumerate(HEADER):
            sheet.cell(row=int(1), column=int(column) + 1).value = title_object

        for row, data_object in enumerate(data):
            for column, sub_obj in enumerate(data_object):
                sheet.cell(row=int(row) + 2, column=int(column) + 1).value = sub_obj
    work_book.save(filename=str(OUTPUT_FILE_PATH))


def formating_data(data: list, params: dict):
    list_data = []
    if params["type"] == "private":
        for row in data:
            list_data.append(
                [
                    "-".join(row["date"]),
                    row["id"],
                    row["name"],
                    row["email"],
                    row["phone_number"],
                    row["subject"],
                    ":".join(row["start_t"]),
                    row["end_time"],
                    row["link"],
                    row["room"],
                    row["group"],
                    row["event_id"],
                    row["event_session_id"],
                ]
            )
    else:
        for row in data:
            list_data.append(
                [
                    "-".join(row["date"]),
                    row["id"],
                    row["name"],
                    row["subject"],
                    ":".join(row["start_t"]),
                    row["end_time"],
                    row["link"],
                    row["group"],
                ]
            )
    return list_data
